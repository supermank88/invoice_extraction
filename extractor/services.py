"""
Invoice extraction service - extracts data from PDF invoices
using DeepSeek API and outputs in Excel schema format.
"""
import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill
from openai import OpenAI

# Schema path
SCHEMA_CSV_PATH = Path(__file__).resolve().parent.parent / "schema_columns.csv"


def _load_schema_columns() -> list[str]:
    """Load column names from schema_columns.csv"""
    with open(SCHEMA_CSV_PATH, encoding="utf-8") as f:
        reader = csv.reader(f)
        row = next(reader)
        return [c.strip() for c in row if c.strip()]


EXCEL_SCHEMA_COLUMNS = _load_schema_columns()


def pdf_to_text(pdf_path: str | Path) -> str:
    """Convert PDF to plain text using pdfplumber."""
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text() or ""
            full_text += "\n"
    return full_text.strip()


def _parse_date(date_str: str) -> datetime | None:
    """Parse date using explicit format strings (no regex)."""
    if not date_str:
        return None
    s = str(date_str).strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        pass
    s = s.replace("Z", "").replace("+00:00", "")
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
        "%d-%b-%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%b %d, %Y",
        "%B %d, %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_unmatched_items_sum(text: str) -> float:
    """Parse amounts from Unmatched Items text (format: 'Desc: 123.45; Desc2: 67.89') and return sum."""
    if not text or not isinstance(text, str):
        return 0.0
    # Match numbers that may have $ or commas, typically after : or at end of segments
    parts = re.split(r"[,;]|\s+(?:and|&)\s+", text)
    total = 0.0
    for part in parts:
        # Look for number pattern (optional $, digits, optional .cents)
        m = re.search(r"\$?\s*([\d,]+(?:\.\d{1,2})?)", part)
        if m:
            try:
                total += float(m.group(1).replace(",", ""))
            except (ValueError, TypeError):
                pass
    return round(total, 2)


def _parse_deepseek_response(response_text: str) -> dict:
    """Parse JSON from DeepSeek response, handling markdown code blocks."""
    text = response_text.strip()
    if "```json" in text:
        start = text.find("```json") + 7
        end = text.find("```", start)
        text = text[start:end] if end > 0 else text[start:]
    elif "```" in text:
        start = text.find("```") + 3
        end = text.find("```", start)
        text = text[start:end] if end > 0 else text[start:]
    return json.loads(text)


def _normalize_extracted_data(raw: dict) -> dict:
    """Normalize DeepSeek output to match schema: correct types, fill missing columns."""
    result = {}
    for col in EXCEL_SCHEMA_COLUMNS:
        val = raw.get(col)
        if val is None or val == "":
            if col == "Bill To":
                result[col] = ""
            elif col in ("INV#", "Date", "Reference"):
                result[col] = None
            else:
                result[col] = 0.0
        elif col == "INV#":
            result[col] = int(val) if isinstance(val, (int, float)) else int(float(str(val)))
        elif col == "Date":
            if isinstance(val, datetime):
                result[col] = val
            elif isinstance(val, str) and val:
                result[col] = _parse_date(val)
            else:
                result[col] = None
        elif col == "Bill To":
            result[col] = str(val).strip() if val else ""
        elif col == "Reference":
            result[col] = str(val).strip() if val else None
        elif col == "Unmatched Items":
            result[col] = str(val).strip() if val else ""
        elif col == "Unmatched Items Value":
            try:
                result[col] = float(val) if isinstance(val, (int, float)) else float(str(val).replace(",", "").replace("$", "").strip() or 0)
            except (ValueError, TypeError):
                result[col] = 0.0
        elif col in ("Subtotal",) or col in EXCEL_SCHEMA_COLUMNS[4:-1]:  # Fee columns
            try:
                if isinstance(val, (int, float)):
                    result[col] = float(val)
                else:
                    result[col] = float(str(val).replace(",", "").replace("$", "").strip())
            except (ValueError, TypeError):
                result[col] = 0.0
        else:
            result[col] = val
    # Fallback: if Unmatched Items has content but Unmatched Items Value is 0, parse sum from text
    unmatched_text = result.get("Unmatched Items", "") or ""
    unmatched_val = result.get("Unmatched Items Value", 0.0) or 0.0
    if unmatched_text and unmatched_val == 0.0:
        parsed_sum = _parse_unmatched_items_sum(unmatched_text)
        if parsed_sum > 0:
            result["Unmatched Items Value"] = parsed_sum
    return result


def extract_invoice_from_pdf(pdf_path: str | Path) -> dict:
    """
    Extract invoice data from PDF: convert to text, then use DeepSeek API to parse.
    Returns dict with keys matching schema_columns.csv.
    """
    return _extract_invoice_internal(pdf_path, use_advanced_rules=False)


def extract_invoice_from_pdf_advanced(pdf_path: str | Path) -> dict:
    """
    Extract invoice data using updated DeepSeek rules. Same flow as extract_invoice_from_pdf
    but with stricter mapping rules to avoid one line item populating multiple columns.
    """
    return _extract_invoice_internal(pdf_path, use_advanced_rules=True)


def _extract_invoice_internal(pdf_path: str | Path, *, use_advanced_rules: bool = False) -> dict:
    """Internal extraction: PDF -> text -> DeepSeek -> normalized dict."""
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        raise ValueError("DEEPSEEK_API_KEY environment variable is not set")

    text = pdf_to_text(pdf_path)
    if not text:
        raise ValueError("Could not extract any text from the PDF")

    columns_str = ", ".join(EXCEL_SCHEMA_COLUMNS)
    base_rules = """- INV#: invoice number (integer)
- Date: invoice date in YYYY-MM-DD format
- Bill To: company/customer name being billed (e.g. "AIO INTERNATIONAL"). Use ONLY the company name, never include the street address (e.g. "6955 136th ST 2A")
- Reference: the invoice reference number, order ID, or reference code (e.g. "USA SEAFOOD - A1000930", "1967"). Never put the Bill To address or street address in Reference
- For fee columns (Customs Duties, Customs Clearance, etc.): use the numeric amount only (e.g. 45.00), 0 if not present
- Map line items to the best matching column (e.g. "CUSTOMS CLEARANCE FEE" -> Customs Clearance, "CUSTOMS ENTRY DUTY & FEE" -> Customs Duties)
- Unmatched Items: any invoice line item that does NOT map to any of the schema columns. Format as "Description: Amount" for each item, separated by "; ". Example: "Misc Service Fee: 25.00; Handling Charge: 10.50". Use empty string "" if all items matched.
- Unmatched Items Value: numeric sum of all amounts from Unmatched Items. Must equal the total of amounts listed in Unmatched Items. Use 0 if no unmatched items.
- Subtotal: total amount from invoice
- Return only valid JSON, no other text"""

    advanced_rules = """
- CRITICAL: Each invoice line item must map to EXACTLY ONE column. Never split one line item across multiple columns.
- Annual Bond vs Annual Customs Bond: These are different. "Annual Bond" (without Customs) -> Annual Bond only. "Annual Customs Bond" or "Customs Bond" (annual) -> Annual Customs Bond only. "Annual Bond 2024-2025" or "Annual Bond2024-2025" -> use Annual Bond2024-2025 only. Do not duplicate amounts in both Annual Bond and Annual Customs Bond.
- Similar rule for all fee columns: one line, one column. Choose the single best-matching column.
- If unsure between two similar columns, pick the one that matches the exact wording on the invoice.
- Unmatched Items: List every line item that does NOT fit any schema column. Format: "Description: Amount" per item, "; " between items. Empty string if all items matched. This helps identify items causing validation mismatches.
- Unmatched Items Value: Sum of all amounts from Unmatched Items. Must match the total of amounts in Unmatched Items. Use 0 if no unmatched items. This is included in Actual Subtotal calculation.
"""

    rules = base_rules + (advanced_rules if use_advanced_rules else "")

    prompt = f"""Extract invoice data from the following text and return a JSON object with exactly these keys, mapping invoice fields to the appropriate columns.

Schema columns (use these exact key names in your JSON):
{columns_str}

Rules:
{rules}

Invoice text:
---
{text}
---
"""

    system_content = "You extract invoice data and return only valid JSON. Bill To = company name only (no address). Reference = ref/order ID only (never an address). No explanations."
    if use_advanced_rules:
        system_content += " Each line item maps to exactly ONE column. Never put the same amount in two columns."

    client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
    response = client.chat.completions.create(
        model="deepseek-chat",
        messages=[
            {"role": "system", "content": system_content},
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )

    content = response.choices[0].message.content
    raw = _parse_deepseek_response(content)
    return _normalize_extracted_data(raw)


def extracted_to_excel(data: dict, output_path: str | Path) -> Path:
    """Write extracted data to Excel file matching schema."""
    return extracted_to_excel_batch([data], output_path)


def extracted_to_excel_batch(
    data_list: list[dict], output_path: str | Path, *, columns: list[str] | None = None
) -> Path:
    """Write multiple extracted records to Excel file. Uses columns or EXCEL_SCHEMA_COLUMNS."""
    headers = columns if columns is not None else EXCEL_SCHEMA_COLUMNS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, data in enumerate(data_list, start=2):
        for col_idx, col_name in enumerate(headers, 1):
            val = data.get(col_name)
            if isinstance(val, datetime):
                val = val.date() if hasattr(val, "date") else val
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_path)
    return Path(output_path)


def extracted_to_csv_compatible_dict(data: dict) -> dict:
    """Convert extracted data to CSV-compatible format (all values as strings)."""
    result = {}
    for k, v in data.items():
        if isinstance(v, datetime):
            result[k] = v.strftime("%Y-%m-%d") if v else ""
        elif v is None:
            result[k] = ""
        else:
            result[k] = str(v)
    return result
