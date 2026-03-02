#!/usr/bin/env python3
"""
Validator script for invoices_extracted.xlsx.
Adds "Actual Subtotal" (sum of fee columns) and "Validate" (Yes/No) columns.
"""
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

# Paths
BASE_DIR = Path(__file__).resolve().parent
SCHEMA_CSV = BASE_DIR / "schema_columns.csv"
DEFAULT_INPUT = BASE_DIR / "invoices_extracted.xlsx"

# Columns to exclude from Actual Subtotal calculation
EXCLUDE_FROM_SUM = {"INV#", "Date", "Bill To", "Reference", "Subtotal"}


def load_schema_columns() -> list[str]:
    """Load column names from schema_columns.csv."""
    with open(SCHEMA_CSV, encoding="utf-8") as f:
        reader = csv.reader(f)
        row = next(reader)
        return [c.strip() for c in row if c.strip()]


def safe_float(val) -> float:
    """Convert value to float, return 0.0 on failure."""
    if val is None or val == "":
        return 0.0
    try:
        s = str(val).replace(",", "").replace("$", "").strip()
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


def main() -> int:
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: {input_path} not found.")
        return 1

    columns = load_schema_columns()
    sum_columns = [c for c in columns if c not in EXCLUDE_FROM_SUM]

    output_path = input_path.parent / (
        "invoices_validated.xlsx" if input_path == DEFAULT_INPUT else f"{input_path.stem}_validated.xlsx"
    )
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Find column indices by header (row 1)
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_index = {str(h): i + 1 for i, h in enumerate(header_row) if h is not None}

    # Insert "Other" column before Subtotal
    subtotal_col = col_index.get("Subtotal")
    if subtotal_col is None:
        print("Error: Subtotal column not found.")
        return 1

    ws.insert_cols(subtotal_col)
    other_col = subtotal_col
    subtotal_col_read = subtotal_col + 1  # Subtotal shifted right
    actual_col = subtotal_col + 2
    validate_col = subtotal_col + 3

    # Insert headers: Other, Actual Subtotal, Validate (Subtotal stays in place)
    ws.cell(row=1, column=other_col, value="Other")
    ws.cell(row=1, column=actual_col, value="Actual Subtotal")
    ws.cell(row=1, column=validate_col, value="Validate")

    # Style new headers like existing ones
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in (other_col, actual_col, validate_col):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font

    yes_count = 0
    max_row = ws.max_row

    for row_num in range(2, max_row + 1):
        # Fee sum: sum of fee columns only
        fee_sum = 0.0
        for col_name in sum_columns:
            idx = col_index.get(col_name)
            if idx is not None:
                cell_val = ws.cell(row=row_num, column=idx).value
                fee_sum += safe_float(cell_val)

        # Get original Subtotal (now shifted right)
        subtotal_val = ws.cell(row=row_num, column=subtotal_col_read).value
        subtotal_float = safe_float(subtotal_val)

        # Other = Subtotal - Fee sum (unclassified/misc amount)
        other_val = round(subtotal_float - fee_sum, 2)

        # Actual Subtotal = Fee sum + Other (includes Other)
        actual_subtotal = fee_sum + other_val

        # Validate: Yes if Actual Subtotal equals Subtotal
        is_valid = abs(actual_subtotal - subtotal_float) < 0.01
        validate_str = "Yes" if is_valid else "No"
        if is_valid:
            yes_count += 1

        ws.cell(row=row_num, column=other_col, value=other_val)
        ws.cell(row=row_num, column=actual_col, value=round(actual_subtotal, 2))
        ws.cell(row=row_num, column=validate_col, value=validate_str)

    wb.save(output_path)
    print(f"Saved to {output_path}")
    print(f"Total Validate 'Yes': {yes_count} / {max_row - 1} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
